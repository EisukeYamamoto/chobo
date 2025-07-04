import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
import pandas as pd
from datetime import datetime
from tkcalendar import DateEntry

EXCEL_FILE = "database.xlsx"

def show_transaction_window():
    def load_accounts():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["口座一覧"]
        accounts = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            display_name = f"{row[1]}（{row[3]}）"
            accounts[display_name] = row[0]
        wb.close()
        return accounts

    def get_initial_balance(account_name):
        account_id = accounts[account_name]
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["口座一覧"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == account_id:
                wb.close()
                return float(row[2])
        wb.close()
        return 0.0

    def search_transactions():
        try:
            start = datetime.strptime(start_entry.get(), "%Y-%m-%d")
            end = datetime.strptime(end_entry.get(), "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("日付エラー", "日付は YYYY-MM-DD 形式で入力してください")
            return

        account_name = account_combo.get()
        if not account_name:
            messagebox.showerror("選択エラー", "口座を選択してください")
            return

        account_id = accounts[account_name]
        initial_balance = get_initial_balance(account_name)

        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["取引履歴"]

        raw_records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row) + [None] * (6 - len(row))
            date_str, acc_id, summary, deposit, withdrawal, writer = row
            if acc_id != account_id:
                continue
            try:
                tx_date = datetime.strptime(str(date_str), "%Y-%m-%d")
            except Exception:
                continue
            if start <= tx_date <= end:
                raw_records.append({
                    "date": tx_date,
                    "summary": summary,
                    "deposit": float(deposit or 0),
                    "withdrawal": float(withdrawal or 0),
                    "writer": writer or ""
                })
        wb.close()

        raw_records.sort(key=lambda r: r["date"])
        current_balance = initial_balance
        records = []
        total_deposit = 0
        total_withdrawal = 0

        for r in raw_records:
            current_balance += r["deposit"] - r["withdrawal"]
            records.append([
                r["date"].strftime("%Y-%m-%d"),
                r["summary"],
                f"{r['deposit']:,.0f}" if r["deposit"] else "",
                f"{r['withdrawal']:,.0f}" if r["withdrawal"] else "",
                f"{current_balance:,.0f}",
                r["writer"]
            ])
            total_deposit += r["deposit"]
            total_withdrawal += r["withdrawal"]

        for row in tree.get_children():
            tree.delete(row)
        for r in records:
            tree.insert("", tk.END, values=r)

        nonlocal current_results, final_balance, export_account, export_start, export_end
        current_results = records
        final_balance = current_balance
        export_account = account_name
        export_start = start_entry.get()
        export_end = end_entry.get()

        deposit_text.set(f"合計預入：{total_deposit:,.0f} 円")
        withdrawal_text.set(f"合計引出：{total_withdrawal:,.0f} 円")
        balance_text.set(f"残高：{current_balance:,.0f} 円")

    def export_to_excel():
        if not current_results:
            messagebox.showinfo("出力なし", "出力するデータがありません")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        df = pd.DataFrame(current_results, columns=["日付", "摘要", "預入", "引出", "残高", "記入者"])
        deposit_total = sum(float(r[2].replace(',', '')) for r in current_results if r[2] != "")
        withdrawal_total = sum(float(r[3].replace(',', '')) for r in current_results if r[3] != "")

        df_summary = pd.DataFrame([
            {"日付": "", "摘要": "", "預入": "", "引出": "", "残高": "", "記入者": ""},
            {"日付": "合計預入", "摘要": "", "預入": deposit_total, "引出": "", "残高": "", "記入者": ""},
            {"日付": "合計引出", "摘要": "", "預入": "", "引出": withdrawal_total, "残高": "", "記入者": ""},
            {"日付": "残高", "摘要": "", "預入": final_balance, "引出": "", "残高": "", "記入者": ""}
        ])

        df_out = pd.concat([df, df_summary], ignore_index=True)

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_out.to_excel(writer, sheet_name="取引履歴", startrow=2, index=False)
            wb = writer.book
            ws = writer.sheets["取引履歴"]

            # ヘッダー書き込み
            ws.cell(row=1, column=1, value=f"{export_account}（{export_start} ～ {export_end}）")

            # 列幅自動調整 & wrap text 設定
            from openpyxl.utils import get_column_letter
            for i, col in enumerate(df_out.columns, start=1):
                max_length = max(df_out[col].astype(str).map(len).max(), len(col)) + 2
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = max(10, min(max_length, 40))
                if col == "摘要":
                    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=i, max_col=i):
                        for cell in row:
                            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                else:
                    ws.column_dimensions[col_letter].width = max(10, min(max_length, 30))
                    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=i, max_col=i):
                        for cell in row:
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0'

            # ページ設定：A4縦、すべてを1ページに収める
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0

        messagebox.showinfo("出力完了", f"{file_path} に出力しました")

    accounts = load_accounts()
    current_results = []
    final_balance = 0.0
    export_account = ""
    export_start = ""
    export_end = ""

    deposit_text = tk.StringVar(value="合計預入：―")
    withdrawal_text = tk.StringVar(value="合計引出：―")
    balance_text = tk.StringVar(value="残高：―")

    root = tk.Toplevel()
    root.title("取引履歴の参照と出力")
    root.geometry("900x700")
    root.configure(bg="#f2f2f7")

    font_label = ("Arial", 13)
    font_entry = ("Arial", 13)
    font_button = ("Arial", 13)
    font_tree = ("Arial", 12)

    # 入力部分
    tk.Label(root, text="口座", font=font_label, bg="#f2f2f7").grid(row=0, column=0, sticky="e", padx=10, pady=5)
    account_combo = ttk.Combobox(root, values=list(accounts.keys()), state="readonly", font=font_entry)
    account_combo.grid(row=0, column=1, padx=10, pady=5, sticky="ew")

    tk.Label(root, text="開始日", font=font_label, bg="#f2f2f7").grid(row=1, column=0, sticky="e", padx=10, pady=5)
    start_entry = DateEntry(root, date_pattern='yyyy-mm-dd', font=font_entry)
    start_entry.grid(row=1, column=1, padx=10, pady=5, sticky="ew")

    tk.Label(root, text="終了日", font=font_label, bg="#f2f2f7").grid(row=2, column=0, sticky="e", padx=10, pady=5)
    end_entry = DateEntry(root, date_pattern='yyyy-mm-dd', font=font_entry)
    end_entry.grid(row=2, column=1, padx=10, pady=5, sticky="ew")

    tk.Button(root, text="検索", command=search_transactions, font=font_button, bg="#4CAF50", fg="white", width=20).grid(row=3, column=0, columnspan=2, pady=10)

    # Treeview + スクロールバー
    tree_frame = tk.Frame(root)
    tree_frame.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
    root.grid_rowconfigure(4, weight=1)
    root.grid_columnconfigure(1, weight=1)

    tree = ttk.Treeview(tree_frame, columns=("日付", "摘要", "預入", "引出", "残高", "記入者"), show="headings", height=15)
    style = ttk.Style()
    style.configure("Treeview.Heading", font=font_label)
    style.configure("Treeview", font=font_tree, rowheight=28)

    for col in ("日付", "摘要", "預入", "引出", "残高", "記入者"):
        tree.heading(col, text=col)
        if col in ("預入", "引出", "残高"):
            tree.column(col, width=120, anchor="e")
        else:
            tree.column(col, width=120, anchor="center")

    tree.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    scrollbar.pack(side="right", fill="y")
    tree.configure(yscrollcommand=scrollbar.set)

    # 集計表示
    tk.Label(root, textvariable=deposit_text, font=font_label, bg="#f2f2f7").grid(row=5, column=0, columnspan=2, sticky="n", pady=2)
    tk.Label(root, textvariable=withdrawal_text, font=font_label, bg="#f2f2f7").grid(row=6, column=0, columnspan=2, sticky="n", pady=2)
    tk.Label(root, textvariable=balance_text, font=("Arial", 14, "bold"), bg="#f2f2f7").grid(row=7, column=0, columnspan=2, sticky="n", pady=5)

    # ボタン群
    tk.Button(root, text="Excelに出力", command=export_to_excel, font=font_button, bg="#4CAF50", fg="white", width=20).grid(row=8, column=0, columnspan=2, pady=10)
    tk.Button(root, text="閉じる", command=root.destroy, font=font_button, bg="#f44336", fg="white", width=20).grid(row=9, column=0, columnspan=2, pady=10)
