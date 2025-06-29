import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from datetime import datetime

EXCEL_FILE = "database.xlsx"

def show_edit_transaction_window():
    def load_accounts():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["口座一覧"]
        mapping = {}
        reverse = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            acc_id, acc_name, *_ = row
            acc_type = row[3] if len(row) >= 4 else "普通"
            full_name = f"{acc_name}（{acc_type}）"
            mapping[acc_id] = full_name
            reverse[full_name] = acc_id
        wb.close()
        return mapping, reverse

    def load_transactions():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["取引履歴"]
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            date, acc_id, summary, deposit, withdrawal, writer = row + (None,) * (6 - len(row))
            acc_display = account_map.get(acc_id, f"未登録({acc_id})")
            records.append((date, acc_display, summary, deposit, withdrawal, writer))
        wb.close()
        return records

    def on_select(event):
        selected = tree.focus()
        if not selected:
            return
        values = tree.item(selected, "values")
        date_entry.delete(0, tk.END)
        summary_entry.delete(0, tk.END)
        deposit_entry.delete(0, tk.END)
        withdrawal_entry.delete(0, tk.END)
        writer_entry.delete(0, tk.END)

        date_entry.insert(0, values[0])
        summary_entry.insert(0, values[2])
        deposit_entry.insert(0, values[3])
        withdrawal_entry.insert(0, values[4])
        writer_entry.insert(0, values[5])

    def update_transaction():
        selected = tree.focus()
        if not selected:
            messagebox.showwarning("選択エラー", "修正する取引を選択してください")
            return

        try:
            row_index = int(tree.item(selected, "tags")[0])
            new_date = date_entry.get().strip()
            new_summary = summary_entry.get().strip()
            new_deposit = deposit_entry.get().strip()
            new_withdrawal = withdrawal_entry.get().strip()
            new_writer = writer_entry.get().strip()

            if not new_date or not new_summary or not new_writer:
                messagebox.showwarning("入力エラー", "日付・摘要・記入者を入力してください")
                return

            if (not new_deposit and not new_withdrawal) or (new_deposit and new_withdrawal):
                messagebox.showwarning("入力エラー", "預入または引出のどちらか一方のみを入力してください")
                return

            if new_deposit:
                new_deposit_val = float(new_deposit)
                if new_deposit_val <= 0:
                    raise ValueError
                new_withdrawal_val = None
            else:
                new_withdrawal_val = float(new_withdrawal)
                if new_withdrawal_val <= 0:
                    raise ValueError
                new_deposit_val = None

            datetime.strptime(new_date, "%Y-%m-%d")

            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb["取引履歴"]
            ws.cell(row=row_index + 2, column=1).value = new_date
            ws.cell(row=row_index + 2, column=3).value = new_summary
            ws.cell(row=row_index + 2, column=4).value = new_deposit_val
            ws.cell(row=row_index + 2, column=5).value = new_withdrawal_val
            ws.cell(row=row_index + 2, column=6).value = new_writer

            wb.save(EXCEL_FILE)
            wb.close()

            messagebox.showinfo("更新成功", "取引内容を更新しました")
            win.destroy()
        except ValueError:
            messagebox.showerror("金額エラー", "金額は正の数値で入力してください")
        except Exception as e:
            messagebox.showerror("更新エラー", f"修正中にエラーが発生しました\n{e}")

    def delete_transaction():
        selected = tree.focus()
        if not selected:
            messagebox.showwarning("選択エラー", "削除する取引を選択してください")
            return
        row_index = int(tree.item(selected, "tags")[0])
        confirm = messagebox.askyesno("確認", "この取引を本当に削除しますか？")
        if not confirm:
            return
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb["取引履歴"]
            ws.delete_rows(row_index + 2)
            wb.save(EXCEL_FILE)
            wb.close()
            messagebox.showinfo("削除完了", "取引を削除しました")
            win.destroy()
        except Exception as e:
            messagebox.showerror("削除エラー", f"削除中にエラーが発生しました\n{e}")

    def colored_button(parent, text, command, bg, fg="white"):
        return tk.Button(parent, text=text, command=command, bg=bg, fg=fg,
                         activebackground=bg, activeforeground=fg,
                         font=("Arial", 14), width=12)

    win = tk.Toplevel()
    win.title("取引履歴の修正・削除")
    win.geometry("1000x700")

    font_main = ("Arial", 14)
    font_tree = ("Arial", 13)

    account_map, reverse_map = load_accounts()
    records = load_transactions()

    style = ttk.Style()
    style.configure("Treeview.Heading", font=font_main)
    style.configure("Treeview", font=font_tree, rowheight=30)

    # Treeview + Scrollbar
    tree_frame = tk.Frame(win)
    tree_frame.pack(pady=10, fill="both", expand=True)

    tree_scroll = tk.Scrollbar(tree_frame)
    tree_scroll.pack(side="right", fill="y")

    tree = ttk.Treeview(tree_frame, columns=("日付", "口座", "摘要", "預入", "引出", "記入者"),
                        show="headings", yscrollcommand=tree_scroll.set, height=10)
    tree_scroll.config(command=tree.yview)

    for col in tree["columns"]:
        tree.heading(col, text=col)
        tree.column(col, width=160, anchor="center")
    tree.pack(fill="both", expand=True)
    tree.bind("<<TreeviewSelect>>", on_select)

    for idx, rec in enumerate(records):
        rec_display = [
            rec[0] or "",
            rec[1] or "",
            rec[2] or "",
            rec[3] if rec[3] is not None else "",
            rec[4] if rec[4] is not None else "",
            rec[5] or ""
        ]
        tree.insert("", "end", values=rec_display, tags=(str(idx),))

    # Entry fields
    form_frame = tk.Frame(win)
    form_frame.pack(pady=10)

    def labeled_entry(row, label_text):
        tk.Label(form_frame, text=label_text, font=font_main).grid(row=row, column=0, sticky="e", padx=5, pady=2)
        entry = tk.Entry(form_frame, font=font_main, width=30)
        entry.grid(row=row, column=1, padx=5, pady=2)
        return entry

    date_entry = labeled_entry(0, "日付")
    summary_entry = labeled_entry(1, "摘要")
    deposit_entry = labeled_entry(2, "預入")
    withdrawal_entry = labeled_entry(3, "引出")
    writer_entry = labeled_entry(4, "記入者")

    # Buttons
    btn_frame = tk.Frame(win)
    btn_frame.pack(pady=20)

    colored_button(btn_frame, "修正する", update_transaction, bg="#4CAF50").pack(side="left", padx=10)
    colored_button(btn_frame, "削除する", delete_transaction, bg="#FF7043").pack(side="left", padx=10)
    colored_button(btn_frame, "閉じる", win.destroy, bg="#F44336").pack(side="left", padx=10)
