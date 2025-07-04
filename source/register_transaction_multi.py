import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from tkcalendar import DateEntry
import openpyxl
import os

EXCEL_FILE = "database.xlsx"

def show_register_multi_window():
    if not os.path.exists(EXCEL_FILE):
        messagebox.showerror("ファイルエラー", f"{EXCEL_FILE} が見つかりません。")
        return

    def load_accounts():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["口座一覧"]
        accounts = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            label = f"{row[1]}（{row[3]}）"
            accounts[label] = row[0]
        wb.close()
        return accounts

    def save_transactions(rows, writer):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["取引履歴"]

        if ws.max_column < 6 or ws.cell(row=1, column=6).value != "記入者":
            ws.cell(row=1, column=6, value="記入者")

        account_dict = load_accounts()
        for row in rows:
            date_str, account_label, summary, deposit, withdrawal = row
            account_id = account_dict.get(account_label)
            if account_id:
                ws.append([date_str, account_id, summary, deposit, withdrawal, writer])
        wb.save(EXCEL_FILE)
        wb.close()

    def register_all():
        rows = []
        for i in range(len(entry_list)):
            date = entry_list[i][0].get()
            account = entry_list[i][1].get()
            summary = entry_list[i][2].get().strip()
            deposit = entry_list[i][3].get().strip()
            withdrawal = entry_list[i][4].get().strip()

            if not (date and account and summary):
                continue

            if (not deposit and not withdrawal):
                continue

            if (deposit and withdrawal):
                messagebox.showwarning("入力エラー", f"{i+1}行目に預入と引出が両方入力されています")
                continue

            try:
                if deposit:
                    dep = int(deposit)
                    if dep <= 0:
                        raise ValueError
                else:
                    dep = ""
                if withdrawal:
                    wd = int(withdrawal)
                    if wd <= 0:
                        raise ValueError
                else:
                    wd = ""
            except ValueError:
                messagebox.showwarning("金額エラー", f"{i+1}行目の金額が正の整数ではありません")
                continue

            try:
                datetime.strptime(date, "%Y-%m-%d")
                dep = float(deposit) if deposit else ""
                wd = float(withdrawal) if withdrawal else ""
                rows.append([date, account, summary, dep, wd])
            except:
                continue

        writer = writer_entry.get().strip()
        if not writer:
            messagebox.showwarning("入力エラー", "記入者名を入力してください")
            return

        if not rows:
            messagebox.showwarning("入力エラー", "有効な行がありません")
            return

        save_transactions(rows, writer)
        messagebox.showinfo("登録完了", f"{len(rows)}件の取引を登録しました")
        win.destroy()

    def back_to_menu():
        win.destroy()

    # ウィンドウ設定
    win = tk.Toplevel()
    win.title("入出金一括登録")
    win.geometry("1200x800")
    win.configure(bg="#f0f0f5")

    font = ("Arial", 14)

    # 見出し行
    headers = ["日付", "口座", "摘要", "預入", "引出"]
    for i, h in enumerate(headers):
        label = tk.Label(win, text=h, font=font, bg="#dfe3ee", padx=5, pady=5)
        label.grid(row=0, column=i, padx=2, pady=2)

    # 入力行
    entry_list = []
    account_dict = load_accounts()
    account_labels = list(account_dict.keys())
    for row in range(20):
        date_entry = DateEntry(win, date_pattern='yyyy-mm-dd', font=font)
        date_entry.set_date(datetime.today())
        date_entry.grid(row=row+1, column=0, padx=1, pady=1)

        account_combo = ttk.Combobox(win, values=account_labels, font=font, state="readonly")
        account_combo.grid(row=row+1, column=1, padx=1, pady=1)

        summary_entry = tk.Entry(win, font=font)
        summary_entry.grid(row=row+1, column=2, padx=1, pady=1)

        deposit_entry = tk.Entry(win, font=font)
        deposit_entry.grid(row=row+1, column=3, padx=1, pady=1)

        withdraw_entry = tk.Entry(win, font=font)
        withdraw_entry.grid(row=row+1, column=4, padx=1, pady=1)

        entry_list.append([date_entry, account_combo, summary_entry, deposit_entry, withdraw_entry])

    # 記入者
    tk.Label(win, text="記入者", font=font, bg="#f0f0f5").grid(row=21, column=0, padx=5, pady=15, sticky="e")
    writer_entry = tk.Entry(win, font=font)
    writer_entry.grid(row=21, column=1, padx=5, pady=15, sticky="w")

    # ボタンフレーム（中央揃え）
    button_frame = tk.Frame(win, bg="#f0f0f5")
    button_frame.grid(row=22, column=0, columnspan=5, pady=30)

    tk.Button(button_frame, text="登録", font=font, command=register_all,
              bg="#4caf50", fg="white", width=15).pack(side="left", padx=30)

    tk.Button(button_frame, text="閉じる", font=font, command=back_to_menu,
              bg="#f44336", fg="white", width=15).pack(side="left", padx=30)
