import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from datetime import datetime
import os
from tkcalendar import DateEntry

EXCEL_FILE = "database.xlsx"

def show_register_window():
    if not os.path.exists(EXCEL_FILE):
        messagebox.showerror("ファイルエラー", f"{EXCEL_FILE} が見つかりません。")
        return

    def load_accounts():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["口座一覧"]
        accounts = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, acc_type = row[1], row[3] if len(row) > 3 else ""
            key = f"{name}（{acc_type}）" if acc_type else name
            accounts[key] = row[0]  # 表示名:口座ID
        wb.close()
        return accounts

    def save_transaction(date_str, account_display, summary, deposit, withdrawal, writer):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        account_id = load_accounts()[account_display]
        ws = wb["取引履歴"]
        if ws.max_column < 6 or ws.cell(row=1, column=6).value != "記入者":
            ws.cell(row=1, column=6, value="記入者")
        ws.append([date_str, account_id, summary, deposit, withdrawal, writer])
        wb.save(EXCEL_FILE)
        wb.close()

    def get_current_balance(account_display):
        account_id = load_accounts()[account_display]
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws1 = wb["口座一覧"]
        ws2 = wb["取引履歴"]
        initial = 0
        for row in ws1.iter_rows(min_row=2, values_only=True):
            if row[0] == account_id:
                initial = float(row[2])
                break
        total_deposit = 0
        total_withdrawal = 0
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[1] == account_id:
                total_deposit += float(row[3] or 0)
                total_withdrawal += float(row[4] or 0)
        wb.close()
        return initial + total_deposit - total_withdrawal

    def register_transaction():
        date_str = date_entry.get()
        account_display = account_combo.get()
        summary = summary_entry.get().strip()
        amount = amount_entry.get().strip()
        writer = writer_entry.get().strip()
        mode = mode_var.get()

        if not (date_str and account_display and summary and amount and writer):
            messagebox.showwarning("入力エラー", "すべての項目を入力してください")
            return

        try:
            input_date = datetime.strptime(date_str, "%Y-%m-%d")
            if input_date > datetime.today():
                messagebox.showwarning("日付エラー", "未来の日付は入力できません")
                return
        except ValueError:
            messagebox.showwarning("日付エラー", "日付の形式が正しくありません")
            return

        try:
            float_amount = float(amount)
            if float_amount <= 0 or not float_amount.is_integer():
                raise ValueError
        except ValueError:
            messagebox.showwarning("金額エラー", "金額は正の整数で入力してください")
            return

        deposit = float_amount if mode == "預入" else ""
        withdrawal = float_amount if mode == "引出" else ""

        try:
            save_transaction(date_str, account_display, summary, deposit, withdrawal, writer)
            current_balance = get_current_balance(account_display)
            messagebox.showinfo("登録完了", f"取引が登録されました\n現在の残高: {current_balance:,.0f} 円")
            summary_entry.delete(0, tk.END)
            amount_entry.delete(0, tk.END)
            writer_entry.delete(0, tk.END)
            summary_entry.focus()
        except Exception as e:
            messagebox.showerror("登録エラー", f"保存中にエラーが発生しました：\n{e}")

    def back_to_menu():
        win.destroy()

    win = tk.Toplevel()
    win.title("入出金登録")
    win.geometry("500x500")
    win.minsize(400, 300)

    for i in range(8):
        win.grid_rowconfigure(i, weight=0)
    win.grid_rowconfigure(7, weight=1)
    win.grid_columnconfigure(0, weight=0)
    win.grid_columnconfigure(1, weight=1)

    font_label = ("Arial", 12)
    font_entry = ("Arial", 12)
    font_button = ("Arial", 12)

    tk.Label(win, text="日付", font=font_label).grid(row=0, column=0, padx=5, pady=5, sticky="e")
    date_entry = DateEntry(win, date_pattern='yyyy-mm-dd', font=font_entry)
    date_entry.set_date(datetime.today())
    date_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

    tk.Label(win, text="口座", font=font_label).grid(row=1, column=0, padx=5, pady=5, sticky="e")
    account_combo = ttk.Combobox(win, state="readonly", font=font_entry)
    account_combo.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
    try:
        account_combo["values"] = list(load_accounts().keys())
    except Exception as e:
        messagebox.showerror("口座読み込みエラー", f"口座一覧の読み込みに失敗しました\n{e}")
        win.destroy()
        return

    tk.Label(win, text="摘要", font=font_label).grid(row=2, column=0, padx=5, pady=5, sticky="e")
    summary_entry = tk.Entry(win, font=font_entry)
    summary_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

    tk.Label(win, text="金額", font=font_label).grid(row=3, column=0, padx=5, pady=5, sticky="e")
    amount_entry = tk.Entry(win, font=font_entry)
    amount_entry.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

    tk.Label(win, text="記入者", font=font_label).grid(row=4, column=0, padx=5, pady=5, sticky="e")
    writer_entry = tk.Entry(win, font=font_entry)
    writer_entry.grid(row=4, column=1, padx=5, pady=5, sticky="ew")

    mode_var = tk.StringVar(value="預入")
    mode_frame = tk.Frame(win)
    mode_frame.grid(row=5, column=0, columnspan=2, pady=10)
    tk.Radiobutton(mode_frame, text="預入", variable=mode_var, value="預入", font=font_entry).pack(side="left", padx=10)
    tk.Radiobutton(mode_frame, text="引出", variable=mode_var, value="引出", font=font_entry).pack(side="left", padx=10)

    tk.Button(win, text="登録", command=register_transaction, font=font_button).grid(row=6, column=0, columnspan=2, pady=10)
    tk.Button(win, text="メニューに戻る", command=back_to_menu, font=font_button).grid(row=7, column=0, columnspan=2, pady=10)

    summary_entry.focus()
