import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

FILE_NAME = "records.xlsx"
SHEET_NAME = "入出金記録"

# Excelファイルを初期化（なければ作成）
def initialize_excel():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        headers = ["日付", "口座名", "摘要", "預入額", "引出額"]
        ws.append(headers)
        wb.save(FILE_NAME)

# Excelに追記
def add_record_to_excel(date, account, desc, deposit, withdraw):
    wb = load_workbook(FILE_NAME)
    ws = wb[SHEET_NAME]
    ws.append([date, account, desc, deposit, withdraw])
    wb.save(FILE_NAME)

# Excelから読み込み
def load_records():
    if not os.path.exists(FILE_NAME):
        return []
    wb = load_workbook(FILE_NAME)
    ws = wb[SHEET_NAME]
    return list(ws.iter_rows(min_row=2, values_only=True))

# GUIアプリ
class FinanceApp:
    def __init__(self, root):
        self.root = root
        root.title("入出金管理アプリ")
        self.create_widgets()
        self.load_table()

    def create_widgets(self):
        # フォーム
        frm_input = ttk.Frame(self.root, padding=10)
        frm_input.pack(fill=tk.X)

        self.var_date = tk.StringVar(value=datetime.today().strftime('%Y-%m-%d'))
        self.var_account = tk.StringVar()
        self.var_desc = tk.StringVar()
        self.var_deposit = tk.StringVar()
        self.var_withdraw = tk.StringVar()

        ttk.Label(frm_input, text="日付").grid(row=0, column=0)
        ttk.Entry(frm_input, textvariable=self.var_date, width=12).grid(row=1, column=0)

        ttk.Label(frm_input, text="口座名").grid(row=0, column=1)
        self.account_combo = ttk.Combobox(frm_input, textvariable=self.var_account, values=["〇〇銀行", "✕✕信金", "△△銀行"], width=15)
        self.account_combo.grid(row=1, column=1)

        ttk.Label(frm_input, text="摘要").grid(row=0, column=2)
        ttk.Entry(frm_input, textvariable=self.var_desc, width=20).grid(row=1, column=2)

        ttk.Label(frm_input, text="預入額").grid(row=0, column=3)
        ttk.Entry(frm_input, textvariable=self.var_deposit, width=10).grid(row=1, column=3)

        ttk.Label(frm_input, text="引出額").grid(row=0, column=4)
        ttk.Entry(frm_input, textvariable=self.var_withdraw, width=10).grid(row=1, column=4)

        ttk.Button(frm_input, text="登録", command=self.register_record).grid(row=1, column=5, padx=5)

        # テーブル
        self.tree = ttk.Treeview(self.root, columns=("日付", "口座名", "摘要", "預入額", "引出額"), show="headings")
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, anchor=tk.CENTER)
        self.tree.pack(fill=tk.BOTH, expand=True, pady=10)

    def register_record(self):
        date = self.var_date.get()
        account = self.var_account.get()
        desc = self.var_desc.get()
        deposit = self.var_deposit.get()
        withdraw = self.var_withdraw.get()

        if not account:
            messagebox.showwarning("入力エラー", "口座名を入力してください")
            return

        try:
            deposit_val = float(deposit) if deposit else 0.0
            withdraw_val = float(withdraw) if withdraw else 0.0
        except ValueError:
            messagebox.showerror("数値エラー", "預入額・引出額には数値を入力してください")
            return

        add_record_to_excel(date, account, desc, deposit_val, withdraw_val)
        self.load_table()
        self.var_desc.set("")
        self.var_deposit.set("")
        self.var_withdraw.set("")

    def load_table(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        records = load_records()
        for rec in records:
            self.tree.insert("", tk.END, values=rec)

if __name__ == "__main__":
    initialize_excel()
    root = tk.Tk()
    app = FinanceApp(root)
    root.mainloop()
