import tkinter as tk
from tkinter import messagebox, ttk
import openpyxl
import os

EXCEL_FILE = "database.xlsx"

def show_add_account_window():
    if not os.path.exists(EXCEL_FILE):
        messagebox.showerror("ファイルエラー", f"{EXCEL_FILE} が見つかりません。")
        return

    def generate_account_id(ws):
        used_ids = [str(row[0]) for row in ws.iter_rows(min_row=2, values_only=True)]
        prefix = "A"
        max_num = 0
        for uid in used_ids:
            if uid.startswith(prefix) and uid[1:].isdigit():
                max_num = max(max_num, int(uid[1:]))
        return f"{prefix}{max_num + 1:03d}"

    def add_account():
        acc_id = id_entry.get().strip()
        acc_name = name_entry.get().strip()
        init_balance = balance_entry.get().strip()
        account_type = type_var.get()

        if not acc_name or not init_balance or not account_type:
            messagebox.showwarning("入力エラー", "すべての項目を入力してください")
            return

        try:
            init_balance = float(init_balance)
        except ValueError:
            messagebox.showwarning("金額エラー", "初期残高は数値で入力してください")
            return

        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb["口座一覧"]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if acc_name == row[1] and account_type == row[3]:
                    messagebox.showerror("重複エラー", "同じ口座名・種別の組み合わせがすでに存在します")
                    wb.close()
                    return

            new_id = generate_account_id(ws)
            ws.append([new_id, acc_name, init_balance, account_type])
            wb.save(EXCEL_FILE)
            wb.close()

            messagebox.showinfo("登録完了", f"{acc_name} を追加しました（ID: {new_id}）")

            id_entry.config(state="normal")
            id_entry.delete(0, tk.END)
            id_entry.insert(0, generate_account_id(openpyxl.load_workbook(EXCEL_FILE)["口座一覧"]))
            id_entry.config(state="readonly")
            name_entry.delete(0, tk.END)
            balance_entry.delete(0, tk.END)
            type_combo.set("")
        except Exception as e:
            messagebox.showerror("エラー", f"登録中にエラーが発生しました\n{e}")

    def colored_button(parent, text, command, bg, fg="white"):
        return tk.Button(parent, text=text, command=command, bg=bg, fg=fg,
                         activebackground=bg, activeforeground=fg,
                         font=("Arial", 14), width=10)

    win = tk.Toplevel()
    win.title("口座の追加")
    win.geometry("400x450")
    win.minsize(300, 300)

    font_label = ("Arial", 14)
    font_entry = ("Arial", 14)

    tk.Label(win, text="口座ID（自動）", font=font_label).pack(pady=(20, 4))
    id_entry = tk.Entry(win, font=font_entry, state="readonly")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        id_entry_val = generate_account_id(wb["口座一覧"])
        wb.close()
    except:
        id_entry_val = "A001"
    id_entry.insert(0, id_entry_val)
    id_entry.pack(pady=2)

    tk.Label(win, text="口座名", font=font_label).pack(pady=(15, 2))
    name_entry = tk.Entry(win, font=font_entry)
    name_entry.pack(pady=2)

    tk.Label(win, text="初期残高", font=font_label).pack(pady=(15, 2))
    balance_entry = tk.Entry(win, font=font_entry)
    balance_entry.pack(pady=2)

    tk.Label(win, text="口座種別", font=font_label).pack(pady=(15, 2))
    type_var = tk.StringVar()
    type_combo = ttk.Combobox(win, textvariable=type_var, state="readonly", font=font_entry)
    type_combo["values"] = ["普通", "定期", "当座"]
    type_combo.pack(pady=2)

    btn_frame = tk.Frame(win)
    btn_frame.pack(pady=30)

    colored_button(btn_frame, "登録", add_account, bg="#4CAF50").pack(side="left", padx=10)
    colored_button(btn_frame, "閉じる", win.destroy, bg="#F44336").pack(side="left", padx=10)
