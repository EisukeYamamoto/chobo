import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
import pandas as pd

EXCEL_FILE = "database.xlsx"

def show_balance_window():
    def load_balances():
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws_accounts = wb["口座一覧"]
            ws_transactions = wb["取引履歴"]

            accounts = {}
            for row in ws_accounts.iter_rows(min_row=2, values_only=True):
                account_id, name, initial, account_type = row[:4]
                display_name = f"{name}（{account_type}）"
                accounts[account_id] = {
                    "name": display_name,
                    "initial": float(initial),
                    "deposit": 0.0,
                    "withdrawal": 0.0
                }

            for row in ws_transactions.iter_rows(min_row=2, values_only=True):
                _, acc_id, _, deposit, withdrawal, *_ = row
                if acc_id in accounts:
                    accounts[acc_id]["deposit"] += float(deposit or 0)
                    accounts[acc_id]["withdrawal"] += float(withdrawal or 0)

            results = []
            total_balance = 0
            for acc in accounts.values():
                current = acc["initial"] + acc["deposit"] - acc["withdrawal"]
                results.append([
                    acc["name"],
                    acc["deposit"],
                    acc["withdrawal"],
                    current
                ])
                total_balance += current
            wb.close()
            return results, total_balance
        except Exception as e:
            messagebox.showerror("エラー", f"残高データの読み込みに失敗しました\n{e}")
            return [], 0

    def show_balances():
        nonlocal current_data, current_total
        records, total = load_balances()
        current_data = records
        current_total = total
        for row in tree.get_children():
            tree.delete(row)
        for r in records:
            tree.insert("", tk.END, values=[
                r[0],
                f"{r[1]:,.0f}",
                f"{r[2]:,.0f}",
                f"{r[3]:,.0f}"
            ])
        total_balance_text.set(f"全口座の残高合計：{total:,.0f} 円")

    def export_to_excel():
        if not current_data:
            messagebox.showinfo("出力なし", "出力するデータがありません")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excelファイル", "*.xlsx")])
        if not file_path:
            return

        df = pd.DataFrame(
            [row[:4] for row in current_data],
            columns=["口座", "合計預入", "合計引出", "現在残高"]
        )
        df_summary = pd.DataFrame([{
            "口座": "全口座の残高合計",
            "合計預入": "",
            "合計引出": "",
            "現在残高": current_total
        }])
        df_out = pd.concat([df, df_summary], ignore_index=True)

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_out.to_excel(writer, sheet_name="残高一覧", startrow=2, index=False)
            wb = writer.book
            ws = writer.sheets["残高一覧"]

            # ヘッダー書き込み
            ws.cell(row=1, column=1, value="全口座の残高一覧")

            # 列幅自動調整 & wrap text 設定（口座列のみ折り返し）
            from openpyxl.utils import get_column_letter
            for i, col in enumerate(df_out.columns, start=1):
                max_length = max(df_out[col].astype(str).map(len).max(), len(col)) + 2
                col_letter = get_column_letter(i)

                if col == "口座":
                    ws.column_dimensions[col_letter].width = max(15, min(max_length, 60))
                    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=i, max_col=i):
                        for cell in row:
                            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                else:
                    ws.column_dimensions[col_letter].width = max(10, min(max_length, 40))
                    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=i, max_col=i):
                        for cell in row:
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0'

            # ページ設定：A4縦、幅内に収める
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0

        messagebox.showinfo("出力完了", f"{file_path} に出力しました")

    def colored_button(parent, text, command, bg, fg="white"):
        return tk.Button(parent, text=text, command=command,
                         bg=bg, fg=fg,
                         activebackground=bg, activeforeground=fg,
                         font=("Arial", 14), width=14)

    balance_win = tk.Toplevel()
    balance_win.title("残高一覧")
    balance_win.geometry("800x600")
    balance_win.minsize(600, 450)

    font_label = ("Arial", 14)
    font_tree = ("Arial", 13)

    current_data = []
    current_total = 0
    total_balance_text = tk.StringVar(value="全口座の残高合計：― 円")

    # ===== 全体ラッパーフレーム =====
    main_frame = tk.Frame(balance_win)
    main_frame.pack(fill="both", expand=True)

    # ===== TreeView + Scrollbar部分 =====
    tree_frame = tk.Frame(main_frame)
    tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

    tree = ttk.Treeview(tree_frame, columns=("口座", "合計預入", "合計引出", "現在残高"), show="headings")
    for col in ("口座", "合計預入", "合計引出", "現在残高"):
        tree.heading(col, text=col)
        if col in ("合計預入", "合計引出", "現在残高"):
            tree.column(col, width=180, anchor="e")
        else:
            tree.column(col, width=180, anchor="center")

    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    tree.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    style = ttk.Style()
    style.configure("Treeview.Heading", font=font_label)
    style.configure("Treeview", font=font_tree, rowheight=28)

    # ===== 残高合計表示 =====
    tk.Label(main_frame, textvariable=total_balance_text, font=("Arial", 14, "bold")).pack(pady=5)

    # ===== ボタン群 =====
    btn_frame = tk.Frame(main_frame)
    btn_frame.pack(pady=10)

    colored_button(btn_frame, "更新", show_balances, bg="#2196F3").pack(side="left", padx=10)
    colored_button(btn_frame, "Excelに出力", export_to_excel, bg="#4CAF50").pack(side="left", padx=10)
    colored_button(btn_frame, "閉じる", balance_win.destroy, bg="#F44336").pack(side="left", padx=10)

    show_balances()
